import pandas as pd
import difflib
import numpy as np
from manga_ocr import MangaOcr
from PIL import ImageGrab, Image
import pyperclip
from loguru import logger
from manga_ocr.run import are_images_identical
import time 
import gradio as gr
import jinja2
import re
from jinja2 import FileSystemLoader, Environment
import glob
from pathlib import Path
from openpyxl import load_workbook
from tkinter import filedialog, Tk
import os

def find_match(s, ref, col2match='jp'):
    matched_lines = difflib.get_close_matches(s, ref[col2match], cutoff=0.5)
    rows = [ref[ref[col2match]==ln] for ln in matched_lines]
    if len(rows)>0:
        return pd.concat(rows,axis=0)
    else:
        return None

def process_results(mocr, img_or_path):
    t0 = time.time()
    text = mocr(img_or_path)
    t1 = time.time()

    logger.info(f'Text recognized in {t1 - t0:0.03f} s: {text}')

    return text

def clean_up(s):
    s = re.sub('<.*?>', '', s)
    s = re.sub(r'\\n',' ', s)
    return s

def clean_up_df(df):
    df.loc[:,'jp'] = df['jp'].apply(clean_up)
    df.loc[:,'eng'] = df['eng'].apply(clean_up)
    
    return df

def parse_text_output(history, env):
    template = env.get_template('text_table.txt')
    return template.render(history=history)

def process_clipboard(game_script_path, ste):
    story = state['story']
    history = state['history']
    
    #load game script on first run
    if story is None:
        load_gamescript(game_script_path, state)        
        story = state['story']
        assert story is not None
    
    last_text = state['last_text']
    try:
        img = ImageGrab.grabclipboard()
        # print(img)
    except OSError as error:
        if "cannot identify image file" in str(error):
            # Pillow error when clipboard hasn't changed since last grab (Linux)
            pass
        elif "target image/png not available" in str(error):
            # Pillow error when clipboard contains text (Linux, X11)
            pass
        else:
            logger.warning('Error while reading from clipboard ({})'.format(error))
    else:
        if isinstance(img, Image.Image):
            '''
            if found a match in game script file, return the corresponding rows in the gamescript
            If no match is found, return the OCR text
            only do recognization if the image is different to save processing power
            '''
            if not are_images_identical(img, state['last_img']):
                print('Processing new image')
                text = process_results(mocr, img)
                matched_line = find_match(text, story)
                if matched_line is not None:
                    jptext = matched_line.iloc[0].jp
                    df_results = matched_line[['jp','eng']]
                    df_results = clean_up_df(df_results)
                else:
                    jptext = text
                    df_results = pd.DataFrame([{'jp':text, 'eng':''}])
                    
                pyperclip.copy(jptext)
                state['last_img'] = img 
                state['last_text']  = text
                
                history.appendleft(df_results)
            else:
                print('Same images. skipping')
            
        html = parse_text_output(history)
        state['history'] = history
        return state, html
        
def load_gamescript(filename):
    #load the gamescript file
    path2load = filename
    print('Loading ', path2load) 
    # check the script content to see how to load it
    wb = load_workbook(path2load, read_only=True)
    if 'Story' in wb.sheetnames:
        story = pd.read_excel(path2load, names=[
            'index', 'cha_jp', 'cha_eng', 'jp', 'eng'], sheet_name='Story')
    else:
        story = pd.read_excel(path2load)

    story = story.fillna('')
    return story
    
def get_dir_and_file(file_path):
    dir_path, file_name = os.path.split(file_path)
    return (dir_path, file_name)


def get_any_file_path(file_path=''):

    current_file_path = file_path

    initial_dir, initial_file = get_dir_and_file(file_path)

    root = Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    file_path = filedialog.askopenfilename(
        initialdir=initial_dir,
        initialfile=initial_file,
    )
    root.destroy()

    if file_path == '':
        file_path = current_file_path

    return file_path